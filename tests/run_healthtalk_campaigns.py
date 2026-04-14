#!/usr/bin/env python3
"""
HealthTalk AI - Doctor Specialty Campaign Runner
Reads HTAI_Campaign_Segments_Doctor_v3.xlsx and runs one Instantly campaign
per specialty sheet:
  - Import leads from Excel
  - Validate emails (LeadMagic + Million Verifier)
  - Generate personalized emails (Claude Sonnet + Haiku review)
  - Sync to Instantly as DRAFT

SAFETY: Campaigns are never activated. No emails are sent.

Usage:
  python -m tests.run_healthtalk_campaigns
  python -m tests.run_healthtalk_campaigns --sheets "A-OBGYN,B-Plastic Surgery"
  python -m tests.run_healthtalk_campaigns --sheets "A-OBGYN" --sender "Alex"
"""

import os
import sys
import time
import argparse
import logging

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from config import INSTANTLY_API_KEY, ANTHROPIC_API_KEY, HAIKU_MODEL

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    datefmt="%H:%M:%S",
    handlers=[logging.StreamHandler(sys.stdout)],
)
logger = logging.getLogger("htai_campaigns")

EXCEL_FILE = r"C:\Users\akash\Downloads\HTAI_Campaign_Segments_Doctor_v3.xlsx"

# ── SAFETY GUARD ─────────────────────────────────────────────────
import campaigns.instantly_client as _instantly_module

def _blocked_activate(self, campaign_id):
    raise RuntimeError("SAFETY GUARD: activate_campaign() is BLOCKED. No emails sent.")

def _blocked_set_sequences(self, campaign_id, sequences):
    raise RuntimeError("SAFETY GUARD: set_campaign_sequences() is BLOCKED.")

_instantly_module.InstantlyClient.activate_campaign      = _blocked_activate
_instantly_module.InstantlyClient.set_campaign_sequences = _blocked_set_sequences


# ── Specialty configs ─────────────────────────────────────────────
# Each entry: sheet_name -> {campaign_name, case_study_key, batch_size}
SPECIALTY_CONFIGS = {
    "A-OBGYN": {
        "campaign": "HTAI - OBGYN & Obstetrics",
        "case_study": "allegiance",
    },
    "B-Plastic Surgery": {
        "campaign": "HTAI - Plastic Surgery",
        "case_study": "jackson_hinds",
    },
    "C-Orthopedics": {
        "campaign": "HTAI - Orthopedics",
        "case_study": "jackson_hinds",
    },
    "D-Dermatology": {
        "campaign": "HTAI - Dermatology",
        "case_study": "jackson_hinds",
    },
    "E-Cardiology": {
        "campaign": "HTAI - Cardiology",
        "case_study": "jackson_hinds",
    },
    "F-Neurology": {
        "campaign": "HTAI - Neurology",
        "case_study": "medcura",
    },
    "G-General Surgery": {
        "campaign": "HTAI - General Surgery",
        "case_study": "medcura",
    },
    "H-Primary Care": {
        "campaign": "HTAI - Primary Care",
        "case_study": "allegiance",
    },
    "I-Psychiatry Psych": {
        "campaign": "HTAI - Psychiatry & Psychology",
        "case_study": "fhcw",
    },
    "J-Social Therapy": {
        "campaign": "HTAI - Social & Therapy",
        "case_study": "medcura",
    },
    "K-Pediatrics": {
        "campaign": "HTAI - Pediatrics",
        "case_study": "jackson_hinds",
    },
    "L-Chiro Rehab": {
        "campaign": "HTAI - Chiropractic & Rehab",
        "case_study": "medcura",
    },
    "M-Nurse Pract": {
        "campaign": "HTAI - Nurse Practitioners",
        "case_study": "fhcw",
    },
    "N-Other Specialty": {
        "campaign": "HTAI - Other Specialties",
        "case_study": "jackson_hinds",
    },
}

# ── Case study content ────────────────────────────────────────────
CASE_STUDIES = {
    "allegiance": {
        "summary": (
            "Allegiance Health Management (rural Louisiana, 9 hospitals, 72 clinics) "
            "used HealthTalk A.I. to run automated Annual Wellness Visit outreach. "
            "In just 6 weeks: 21,395 patients contacted, 1,768 visits scheduled, "
            "1,783 call center hours saved, $373,048 in new revenue, $44,575 in cost savings. "
            "Total impact: $417,623 in 6 weeks."
        ),
    },
    "jackson_hinds": {
        "summary": (
            "Jackson-Hinds Comprehensive Health Center (Mississippi's largest FQHC, 34 locations) "
            "used HealthTalk A.I. to close care gaps across 6 patient cohorts including "
            "well-child visits, adult physicals, OB-GYN Well-Woman visits, and patient re-engagement. "
            "In just over 1 month: 19,534 patients contacted, 896 visits scheduled, "
            "1,628 call center hours saved, $122,977 in projected revenue and cost savings."
        ),
    },
    "medcura": {
        "summary": (
            "MedCura Health (14 community health locations in metro-Atlanta) "
            "used HealthTalk A.I. to automate referral management and loop closure. "
            "Results: 13,981 patients contacted, 53.54% response rate (vs <10% by phone), "
            "9,689 referral loops closed (69.3%), 2,261 staff hours saved — valued at $41,000. "
            "Equivalent of a full-time employee freed for higher-value care coordination."
        ),
    },
    "fhcw": {
        "summary": (
            "Family Health Center of Worcester (FQHC, Worcester MA) "
            "used HealthTalk A.I. to solve after-hours care coverage with on-demand Telehealth. "
            "In the first 2 months: 260 adult on-demand care requests completed (42% billed as "
            "after-hours Telehealth), 19 pediatric care requests (62% billed). "
            "Provider retention improved. Patients seen at any time without leaving home. "
            "No downloads or login credentials needed."
        ),
    },
}

# ── HealthTalk AI brief ───────────────────────────────────────────
HTAI_VALUE_PROP = (
    "HealthTalk A.I. is a HIPAA-compliant, SOC 2 Type 1 certified, KLAS-recognized "
    "AI patient engagement platform that integrates with 90+ EHR systems. "
    "It automates patient outreach, appointment scheduling, care gap closure, "
    "referral management, and after-hours Telehealth — reducing staff burden "
    "while increasing patient encounters and revenue."
)


# ── Helpers ──────────────────────────────────────────────────────

def header(title):
    logger.info("")
    logger.info("=" * 60)
    logger.info(f"  {title}")
    logger.info("=" * 60)


def query_one(sql, params=None):
    from db import get_connection
    conn = get_connection()
    try:
        with conn.cursor() as cur:
            cur.execute(sql, params)
            return cur.fetchone()
    finally:
        conn.close()


def query_all(sql, params=None):
    from db import get_connection
    from psycopg2.extras import RealDictCursor
    conn = get_connection()
    try:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(sql, params)
            return cur.fetchall()
    finally:
        conn.close()


# ── Phase 0: Pre-flight ──────────────────────────────────────────

def phase_0_preflight():
    header("PHASE 0: PRE-FLIGHT CHECKS")
    import requests
    from config import (
        LEADMAGIC_API_KEY, LEADMAGIC_BASE_URL,
        MILLION_VERIFIER_API_KEY, MILLION_VERIFIER_BASE_URL,
    )

    # DB
    try:
        from db import get_connection
        conn = get_connection()
        with conn.cursor() as cur:
            cur.execute("SELECT current_database()")
            db_name = cur.fetchone()[0]
        conn.close()
        logger.info(f"  [PASS] DB connection - {db_name}")
    except Exception as e:
        logger.error(f"  [FAIL] DB: {e}")
        sys.exit(1)

    # Excel file
    if os.path.exists(EXCEL_FILE):
        logger.info(f"  [PASS] Excel file found")
    else:
        logger.error(f"  [FAIL] Excel file not found: {EXCEL_FILE}")
        sys.exit(1)

    # APIs
    for name, fn in [
        ("LeadMagic", lambda: requests.post(
            f"{LEADMAGIC_BASE_URL}/email-validate",
            headers={"X-API-Key": LEADMAGIC_API_KEY, "Content-Type": "application/json"},
            json={"email": "test@google.com"}, timeout=15)),
        ("Instantly", lambda: requests.get(
            "https://api.instantly.ai/api/v2/campaigns",
            headers={"Authorization": f"Bearer {INSTANTLY_API_KEY}"},
            params={"limit": 1}, timeout=15)),
        ("Anthropic", lambda: __import__("anthropic").Anthropic(
            api_key=ANTHROPIC_API_KEY).messages.create(
            model=HAIKU_MODEL, max_tokens=5,
            messages=[{"role": "user", "content": "Hi"}])),
    ]:
        try:
            r = fn()
            status = getattr(r, "status_code", 200)
            ok = status == 200
            logger.info(f"  {'[PASS]' if ok else '[FAIL]'} {name}")
        except Exception as e:
            logger.error(f"  [FAIL] {name}: {e}")

    logger.info("  Pre-flight complete.\n")


# ── Run one specialty campaign ────────────────────────────────────

def run_specialty(sheet_name: str, config: dict, sender_name: str) -> dict:
    """Run full pipeline for one specialty sheet. Returns summary dict."""
    campaign_name = config["campaign"]
    case_study    = CASE_STUDIES[config["case_study"]]

    header(f"CAMPAIGN: {campaign_name}")
    start = time.time()

    from db import create_client, create_campaign, create_campaign_brief
    from ingestion.csv_importer import import_excel_sheet
    from validation.cascade_validator import validate_lead_email
    from generation.email_generator import generate_batch
    from campaigns.campaign_launcher import launch_campaign
    from tracking.budget_guard import get_budget_status

    # 1. Seed DB
    logger.info("  Creating campaign in DB...")
    client_id = create_client(
        "Infinite Labs Digital",
        INSTANTLY_API_KEY,
        ["infinitelabsdigital.com"]
    )
    campaign_id = create_campaign(
        campaign_name, client_id, "healthtalk_ai", "national", "United States"
    )

    # 2. Create campaign brief with case study
    create_campaign_brief(
        campaign_id,
        service_name="HealthTalk A.I. - AI Patient Engagement Platform",
        service_detail=HTAI_VALUE_PROP,
        value_prop=(
            "Practices using HealthTalk A.I. see more patient encounters, fewer no-shows, "
            "reduced staff workload, and measurable revenue growth — all without adding headcount."
        ),
        case_studies=[case_study],
        sender_name=sender_name,
        sender_title="Account Executive",
        cta_type="call",
        cta_detail="Book a 15-minute demo at HealthTalkAI.com",
        custom_notes=(
            f"This is a cold email to a {sheet_name.split('-', 1)[-1].strip()} physician practice owner. "
            "Reference their specialty naturally. Keep emails under 120 words. "
            "Lead with patient engagement pain points, not features."
        ),
    )
    logger.info(f"  Campaign ID: {campaign_id}")

    # 3. Import leads from Excel
    logger.info(f"  Importing leads from sheet '{sheet_name}'...")
    import_stats = import_excel_sheet(EXCEL_FILE, sheet_name, campaign_id, "htai_client_list")
    logger.info(f"  Import: {import_stats}")

    total_leads = query_one(
        "SELECT COUNT(*) FROM leads WHERE campaign_id = %s", (campaign_id,)
    )[0]
    logger.info(f"  Leads in DB: {total_leads}")

    if total_leads == 0:
        logger.warning("  No leads imported. Skipping campaign.")
        return {"campaign": campaign_name, "leads": 0, "send": 0, "generated": 0, "synced": 0}

    # 4. Email validation
    logger.info("  Validating emails...")
    from db import get_connection
    from psycopg2.extras import RealDictCursor

    # Get leads with emails that haven't been validated
    conn = get_connection()
    try:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("""
                SELECT lead_id, email FROM leads
                WHERE campaign_id = %s
                  AND email IS NOT NULL
                  AND email_verdict IS NULL
                ORDER BY ingested_at ASC
            """, (campaign_id,))
            to_validate = cur.fetchall()
    finally:
        conn.close()

    logger.info(f"  Leads to validate: {len(to_validate)}")
    validated = 0
    from concurrent.futures import ThreadPoolExecutor, as_completed

    def _validate_one(lead):
        try:
            validate_lead_email(dict(lead), campaign_id)
            return True
        except Exception as e:
            logger.error(f"  Validation error for lead {lead['lead_id']}: {e}")
            return False

    with ThreadPoolExecutor(max_workers=10) as pool:
        futures = {pool.submit(_validate_one, lead): lead for lead in to_validate}
        for fut in as_completed(futures):
            if fut.result():
                validated += 1
            if validated % 100 == 0 and validated > 0:
                logger.info(f"    Validated {validated}/{len(to_validate)}...")

    verdict_rows = query_all(
        "SELECT email_verdict, COUNT(*) as cnt FROM leads WHERE campaign_id = %s "
        "AND email_verdict IS NOT NULL GROUP BY email_verdict",
        (campaign_id,)
    )
    verdicts = {r["email_verdict"]: r["cnt"] for r in verdict_rows}
    send_count = verdicts.get("SEND", 0)
    logger.info(f"  Verdicts: {verdicts}")

    if send_count == 0:
        logger.warning("  No SEND leads. Skipping email generation.")
        return {
            "campaign": campaign_name, "leads": total_leads,
            "send": 0, "generated": 0, "synced": 0,
            "elapsed": time.time() - start,
        }

    # 5. Email generation
    budget = get_budget_status()
    logger.info(f"  Budget: ${budget['spent']:.2f} / ${budget['budget']:.2f}")
    logger.info(f"  Generating emails for {send_count} SEND leads...")

    gen_stats = generate_batch(campaign_id, batch_size=send_count + 10)
    logger.info(f"  Generated: {gen_stats}")

    # 6. Instantly sync
    logger.info("  Syncing to Instantly (DRAFT)...")
    launch_stats = launch_campaign(campaign_id)
    logger.info(f"  Synced: {launch_stats}")

    instantly_id = query_one(
        "SELECT instantly_campaign_id FROM campaigns WHERE campaign_id = %s",
        (campaign_id,)
    )
    if instantly_id and instantly_id[0]:
        logger.info(f"  Instantly campaign ID: {instantly_id[0]}")

    elapsed = time.time() - start
    logger.info(f"  Campaign done in {elapsed/60:.1f} min")

    return {
        "campaign":    campaign_name,
        "sheet":       sheet_name,
        "campaign_id": campaign_id,
        "instantly_id": instantly_id[0] if instantly_id and instantly_id[0] else None,
        "leads":       total_leads,
        "validated":   validated,
        "send":        send_count,
        "verdicts":    verdicts,
        "generated":   gen_stats.get("generated", 0),
        "errors":      gen_stats.get("errors", 0),
        "synced":      launch_stats.get("synced", 0),
        "elapsed_min": round(elapsed / 60, 1),
    }


# ── Main ─────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Run HealthTalk AI specialty campaigns")
    parser.add_argument("--sheets", default=None,
                        help="Comma-separated sheet names to run. Default: all sheets.")
    parser.add_argument("--sender", default="Alex",
                        help="Sender first name for email signatures. Default: Alex")
    args = parser.parse_args()

    start_all = time.time()

    if args.sheets:
        sheets_to_run = [s.strip() for s in args.sheets.split(",")]
    else:
        sheets_to_run = list(SPECIALTY_CONFIGS.keys())

    print("\n")
    print("=" * 60)
    print("  HealthTalk AI - Doctor Specialty Campaigns")
    print(f"  Sheets: {', '.join(sheets_to_run)}")
    print(f"  Sender: {args.sender}")
    print("  SAFETY: No emails will be sent (DRAFT only)")
    print("=" * 60)

    # Estimate cost
    import openpyxl
    wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True)
    total_est = sum(
        (wb[s].max_row - 2) for s in sheets_to_run if s in wb.sheetnames
    )
    wb.close()
    print(f"\n  Estimated leads: ~{total_est:,}")
    print(f"  Estimated validation cost: ~${total_est * 0.012:.0f}")
    print(f"  Estimated generation cost (50% pass): ~${total_est * 0.5 * 0.02:.0f}")
    print(f"  Estimated total cost: ~${total_est * 0.022:.0f}")
    print()

    phase_0_preflight()

    results = []
    for sheet_name in sheets_to_run:
        if sheet_name not in SPECIALTY_CONFIGS:
            logger.warning(f"Unknown sheet '{sheet_name}', skipping.")
            continue
        try:
            result = run_specialty(sheet_name, SPECIALTY_CONFIGS[sheet_name], args.sender)
            results.append(result)
        except Exception as e:
            logger.error(f"Campaign failed for {sheet_name}: {e}")
            results.append({"campaign": sheet_name, "error": str(e)})

    # Final summary
    elapsed_all = time.time() - start_all
    print("\n")
    print("=" * 60)
    print("  FINAL SUMMARY - All HealthTalk AI Campaigns")
    print("=" * 60)
    total_leads = total_synced = total_gen = 0
    for r in results:
        if "error" in r:
            print(f"  {r['campaign']}: ERROR - {r['error']}")
            continue
        print(f"  {r['campaign']}")
        print(f"    Leads: {r['leads']} | SEND: {r['send']} | "
              f"Generated: {r['generated']} | Synced: {r['synced']} | "
              f"Instantly: {r.get('instantly_id', 'N/A')}")
        total_leads  += r.get("leads", 0)
        total_synced += r.get("synced", 0)
        total_gen    += r.get("generated", 0)

    from tracking.budget_guard import get_budget_status
    budget = get_budget_status()
    print()
    print(f"  Total leads processed: {total_leads:,}")
    print(f"  Total sequences generated: {total_gen:,}")
    print(f"  Total synced to Instantly: {total_synced:,}")
    print(f"  Total API spend: ${budget['spent']:.2f}")
    print(f"  Total time: {elapsed_all/60:.1f} minutes")
    print("=" * 60)
    print()


if __name__ == "__main__":
    main()